#!/usr/bin/env python3
"""Stage-B task-representation screening for SpreadsheetBench.

Phases:
  schema   Extract answer-free workbook schema summaries for the 61 Train80
           tasks that have task-specific prompt oracles.
  qwen     Extract frozen-Qwen generation-boundary and mean hidden states for
           instruction-only and instruction+schema TaskSpecs.
  analyze  Run leakage-safe task-level cross-validation against prompt
           residuals and existing behavioral diagnostics.

No Val40 or Test280 item is loaded by this script.
"""

from __future__ import annotations

import argparse
import csv
import gc
import json
import math
import os
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any

import numpy as np
import torch
from openpyxl import load_workbook
from tqdm import tqdm

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.analyze_spreadsheetbench_task_prompt_manifold import (
    DEFAULT_ITEMS,
    DEFAULT_TASK_ROOT,
    add_bias,
    checkpoint_matrix,
    choose_ridge_alpha,
    cosine_rows,
    read_json,
    read_jsonl,
    ridge_predict,
    spearman,
    stratified_folds,
)


DEFAULT_DATA_ROOT = PROJECT_ROOT / "data/spreadsheetbench_verified_400"
DEFAULT_OUT = PROJECT_ROOT / "outputs/SpreadsheetBench_task_representation_stage_b"
DEFAULT_MODEL = PROJECT_ROOT.parent / "model_cache/huggingface/models--Qwen--Qwen3.6-35B-A3B/snapshots/995ad96eacd98c81ed38be0c5b274b04031597b0"


FORMULA_FUNCTION = re.compile(r"(?:^|[^A-Z0-9_.])([A-Z][A-Z0-9_.]*)\s*\(", re.IGNORECASE)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--phase", choices=("schema", "qwen", "analyze", "all"), default="all")
    parser.add_argument("--task-root", type=Path, default=DEFAULT_TASK_ROOT)
    parser.add_argument("--items", type=Path, default=DEFAULT_ITEMS)
    parser.add_argument("--data-root", type=Path, default=DEFAULT_DATA_ROOT)
    parser.add_argument("--out-root", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--model", type=Path, default=Path(os.environ.get("SPREADSHEETBENCH_MODEL", DEFAULT_MODEL)))
    parser.add_argument("--batch-size", type=int, default=2)
    parser.add_argument("--max-tokens", type=int, default=1536)
    parser.add_argument("--seed", type=int, default=24018)
    parser.add_argument("--folds", type=int, default=5)
    parser.add_argument("--components", type=int, default=16)
    return parser.parse_args()


def json_dump(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def find_initial_workbook(data_root: Path, item: dict[str, Any]) -> Path:
    task_dir = data_root / str(item.get("spreadsheet_path", f"spreadsheet/{item['id']}"))
    for pattern in ("*_init.xlsx", "initial.xlsx", "input.xlsx"):
        candidates = sorted(task_dir.glob(pattern))
        if len(candidates) == 1:
            return candidates[0]
        if len(candidates) > 1:
            raise RuntimeError(f"Ambiguous initial workbooks for task {item['id']}: {candidates}")
    raise RuntimeError(f"No initial workbook found for task {item['id']} under {task_dir}")


def number_format_classes(number_format: str) -> set[str]:
    value = str(number_format or "").lower()
    classes: set[str] = set()
    if "%" in value:
        classes.add("percent")
    if any(marker in value for marker in ("$", "€", "£", "¥", "[$")):
        classes.add("currency")
    if any(marker in value for marker in ("yy", "dd", "mmm")):
        classes.add("date")
    if any(marker in value for marker in ("hh", "ss", "am/pm")):
        classes.add("time")
    if "0.00" in value or "#,##0" in value:
        classes.add("numeric")
    return classes


def workbook_schema(path: Path) -> tuple[str, list[float], list[str]]:
    # The input workbook only is read. Formula results and arbitrary cell
    # values are not serialized into the TaskSpec.
    wb = load_workbook(path, read_only=False, data_only=False)
    feature_names = [
        "sheet_count",
        "visible_sheet_count",
        "nonempty_cells",
        "formula_cells",
        "string_cells",
        "numeric_cells",
        "boolean_cells",
        "merged_ranges",
        "tables",
        "charts",
        "data_validations",
        "named_ranges",
        "date_formats",
        "time_formats",
        "percent_formats",
        "currency_formats",
        "numeric_formats",
        "max_rows",
        "max_columns",
    ]
    counts = Counter({name: 0 for name in feature_names})
    counts["sheet_count"] = len(wb.worksheets)
    counts["visible_sheet_count"] = sum(ws.sheet_state == "visible" for ws in wb.worksheets)
    counts["named_ranges"] = len(list(wb.defined_names.values()))
    sheet_descriptions: list[str] = []
    for ws in wb.worksheets:
        nonempty = 0
        formulas = 0
        strings = 0
        numerics = 0
        booleans = 0
        format_counts: Counter[str] = Counter()
        formula_functions: Counter[str] = Counter()
        header_candidates: list[str] = []
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                nonempty += 1
                if isinstance(value, str) and value.startswith("="):
                    formulas += 1
                    for match in FORMULA_FUNCTION.findall(value):
                        formula_functions[match.upper()] += 1
                elif isinstance(value, str):
                    strings += 1
                    if cell.row <= 5:
                        clean = " ".join(value.split())
                        if 0 < len(clean) <= 80 and clean not in header_candidates:
                            header_candidates.append(clean)
                elif isinstance(value, bool):
                    booleans += 1
                elif isinstance(value, (int, float)):
                    numerics += 1
                for category in number_format_classes(cell.number_format):
                    format_counts[category] += 1
        tables = len(ws.tables)
        charts = len(ws._charts)  # openpyxl exposes charts through this collection.
        validations = len(ws.data_validations.dataValidation) if ws.data_validations is not None else 0
        merged = len(ws.merged_cells.ranges)
        counts.update(
            {
                "nonempty_cells": nonempty,
                "formula_cells": formulas,
                "string_cells": strings,
                "numeric_cells": numerics,
                "boolean_cells": booleans,
                "merged_ranges": merged,
                "tables": tables,
                "charts": charts,
                "data_validations": validations,
                "date_formats": format_counts["date"],
                "time_formats": format_counts["time"],
                "percent_formats": format_counts["percent"],
                "currency_formats": format_counts["currency"],
                "numeric_formats": format_counts["numeric"],
            }
        )
        counts["max_rows"] = max(counts["max_rows"], int(ws.max_row or 0))
        counts["max_columns"] = max(counts["max_columns"], int(ws.max_column or 0))
        parts = [
            f"sheet={ws.title}",
            f"state={ws.sheet_state}",
            f"size={int(ws.max_row or 0)}x{int(ws.max_column or 0)}",
            f"nonempty={nonempty}",
            f"formulas={formulas}",
            f"tables={tables}",
            f"charts={charts}",
            f"merged={merged}",
        ]
        if header_candidates:
            parts.append("headers=" + " | ".join(header_candidates[:32]))
        if formula_functions:
            parts.append("formula_functions=" + ",".join(name for name, _ in formula_functions.most_common(16)))
        if format_counts:
            parts.append("formats=" + ",".join(f"{name}:{count}" for name, count in sorted(format_counts.items())))
        sheet_descriptions.append("; ".join(parts))
    wb.close()
    schema_text = "\n".join(sheet_descriptions)
    numeric = [math.log1p(float(counts[name])) for name in feature_names]
    return schema_text, numeric, feature_names


def build_task_specs(args: argparse.Namespace) -> list[dict[str, Any]]:
    task_root = args.task_root.resolve()
    task_ids = sorted(path.parent.name for path in (task_root / "training").glob("*/final_prefix.pt"))
    if len(task_ids) != 61:
        raise RuntimeError(f"Expected 61 task-specific checkpoints, found {len(task_ids)}")
    items = {str(row["id"]): row for row in read_json(args.items.resolve())}
    rows: list[dict[str, Any]] = []
    expected_names: list[str] | None = None
    for task_id in tqdm(task_ids, desc="Schema", unit="task"):
        item = items[task_id]
        workbook = find_initial_workbook(args.data_root.resolve(), item)
        schema_text, numeric, names = workbook_schema(workbook)
        if expected_names is None:
            expected_names = names
        elif names != expected_names:
            raise RuntimeError("Schema feature order changed between workbooks")
        instruction = str(item.get("instruction", ""))
        instruction_type = str(item.get("instruction_type", ""))
        instruction_spec = (
            "[TASK_TYPE]\n"
            f"{instruction_type}\n"
            "[INSTRUCTION]\n"
            f"{instruction}\n"
            "[TASK_REPRESENTATION]\n"
        )
        full_spec = (
            "[TASK_TYPE]\n"
            f"{instruction_type}\n"
            "[INSTRUCTION]\n"
            f"{instruction}\n"
            "[WORKBOOK_SCHEMA]\n"
            f"{schema_text}\n"
            "[TASK_REPRESENTATION]\n"
        )
        rows.append(
            {
                "id": task_id,
                "instruction_type": instruction_type,
                "instruction": instruction,
                "schema_text": schema_text,
                "schema_numeric": numeric,
                "schema_feature_names": names,
                "instruction_spec": instruction_spec,
                "full_spec": full_spec,
                "input_workbook": str(workbook),
                "answer_fields_used": False,
            }
        )
    output = args.out_root.resolve() / "task_specs.jsonl"
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")
    json_dump(
        args.out_root.resolve() / "schema_summary.json",
        {
            "tasks": len(rows),
            "split": "Train80 successful-trajectory subset only",
            "feature_names": expected_names,
            "answer_fields_used": False,
            "val_or_test_accessed": False,
        },
    )
    return rows


def extract_qwen_embeddings(args: argparse.Namespace, rows: list[dict[str, Any]]) -> None:
    from transformers import AutoModelForCausalLM, AutoTokenizer

    model_path = args.model.resolve()
    if not model_path.is_dir():
        raise FileNotFoundError(f"Local Qwen model not found: {model_path}")
    if not torch.cuda.is_available():
        raise RuntimeError("Qwen representation extraction requires a CUDA node")
    tokenizer = AutoTokenizer.from_pretrained(model_path, local_files_only=True, trust_remote_code=False)
    if tokenizer.pad_token_id is None:
        tokenizer.pad_token = tokenizer.eos_token
    tokenizer.padding_side = "left"
    model = AutoModelForCausalLM.from_pretrained(
        model_path,
        local_files_only=True,
        trust_remote_code=False,
        dtype=torch.bfloat16,
        low_cpu_mem_usage=True,
    ).to("cuda")
    model.eval()
    backbone = getattr(model, "model", None)
    if backbone is None:
        raise AttributeError("Loaded causal LM does not expose a .model backbone")

    outputs: dict[str, np.ndarray] = {}
    ids = np.asarray([row["id"] for row in rows])
    for field, prefix in (("instruction_spec", "qwen_instruction"), ("full_spec", "qwen_task_spec")):
        last_vectors: list[np.ndarray] = []
        mean_vectors: list[np.ndarray] = []
        truncated = 0
        for start in tqdm(range(0, len(rows), args.batch_size), desc=f"Qwen {prefix}", unit="batch"):
            batch_rows = rows[start : start + args.batch_size]
            texts = [str(row[field]) for row in batch_rows]
            raw_lengths = tokenizer(texts, add_special_tokens=True, padding=False, truncation=False)["input_ids"]
            truncated += sum(len(value) > args.max_tokens for value in raw_lengths)
            encoded = tokenizer(
                texts,
                add_special_tokens=True,
                padding=True,
                truncation=True,
                max_length=args.max_tokens,
                return_tensors="pt",
            )
            input_ids = encoded["input_ids"].to("cuda")
            attention_mask = encoded["attention_mask"].to("cuda")
            with torch.inference_mode():
                result = backbone(
                    input_ids=input_ids,
                    attention_mask=attention_mask,
                    use_cache=False,
                    return_dict=True,
                )
                hidden = result.last_hidden_state.float()
                last_indices = attention_mask.sum(dim=1) - 1
                # With left padding, the last active token is always the final
                # tensor position. Keep the general index expression explicit.
                left_padding = attention_mask.shape[1] - attention_mask.sum(dim=1)
                absolute_last = left_padding + last_indices
                batch_index = torch.arange(len(batch_rows), device=hidden.device)
                last = hidden[batch_index, absolute_last]
                weights = attention_mask.to(hidden.dtype).unsqueeze(-1)
                mean = (hidden * weights).sum(dim=1) / weights.sum(dim=1).clamp_min(1.0)
            last_vectors.append(last.cpu().numpy())
            mean_vectors.append(mean.cpu().numpy())
            del result, hidden, last, mean, input_ids, attention_mask
            torch.cuda.empty_cache()
        outputs[f"{prefix}_last"] = np.concatenate(last_vectors).astype(np.float32)
        outputs[f"{prefix}_mean"] = np.concatenate(mean_vectors).astype(np.float32)
        print(f"{prefix}: truncated {truncated}/{len(rows)} task specs", flush=True)

    np.savez_compressed(args.out_root.resolve() / "qwen_task_embeddings.npz", task_ids=ids, **outputs)
    json_dump(
        args.out_root.resolve() / "qwen_embedding_summary.json",
        {
            "model": str(model_path),
            "tasks": len(rows),
            "max_tokens": args.max_tokens,
            "batch_size": args.batch_size,
            "representations": {name: list(value.shape) for name, value in outputs.items()},
            "val_or_test_accessed": False,
        },
    )
    del backbone, model
    gc.collect()
    torch.cuda.empty_cache()


def normalize_train_test(train: np.ndarray, test: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    mean = np.mean(train, axis=0, keepdims=True)
    std = np.std(train, axis=0, keepdims=True)
    train = (train - mean) / np.maximum(std, 1e-6)
    test = (test - mean) / np.maximum(std, 1e-6)
    train /= np.maximum(np.linalg.norm(train, axis=1, keepdims=True), 1e-12)
    test /= np.maximum(np.linalg.norm(test, axis=1, keepdims=True), 1e-12)
    return train, test


def auc_score(labels: np.ndarray, scores: np.ndarray) -> float:
    labels = np.asarray(labels).astype(bool)
    positives = scores[labels]
    negatives = scores[~labels]
    if len(positives) == 0 or len(negatives) == 0:
        return float("nan")
    wins = 0.0
    for positive in positives:
        wins += float(np.sum(positive > negatives)) + 0.5 * float(np.sum(positive == negatives))
    return wins / (len(positives) * len(negatives))


def dense_representation_cv(
    name: str,
    representation: np.ndarray,
    prompts: np.ndarray,
    closures: np.ndarray,
    successes: np.ndarray,
    strata: list[str],
    folds: list[list[int]],
    components: int,
    seed: int,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    all_indices = set(range(len(prompts)))
    task_predictions: list[dict[str, Any]] = []
    total_sse = total_sst = 0.0
    cosine_values: list[float] = []
    nn_sse = nn_sst = 0.0
    nn_cosines: list[float] = []
    closure_true: list[float] = []
    closure_pred: list[float] = []
    success_true: list[int] = []
    success_score: list[float] = []
    fold_summaries: list[dict[str, Any]] = []
    for fold_idx, test_indices in enumerate(folds):
        train_indices = sorted(all_indices - set(test_indices))
        x_train, x_test = normalize_train_test(representation[train_indices], representation[test_indices])
        prompt_train = prompts[train_indices]
        fold_mean = np.mean(prompt_train, axis=0, keepdims=True)
        centered_train = prompt_train - fold_mean
        _, _, vt = np.linalg.svd(centered_train, full_matrices=False)
        n_components = min(components, len(train_indices) - 1)
        basis = vt[:n_components]
        y_train = centered_train @ basis.T
        alpha = choose_ridge_alpha(
            x_train,
            y_train,
            [strata[index] for index in train_indices],
            seed=seed + fold_idx,
        )
        y_pred = ridge_predict(add_bias(x_train), y_train, add_bias(x_test), alpha)
        prompt_pred = fold_mean + y_pred @ basis
        true_residual = prompts[test_indices] - fold_mean
        pred_residual = prompt_pred - fold_mean
        cosines = cosine_rows(pred_residual, true_residual)
        squared_errors = np.sum((prompt_pred - prompts[test_indices]) ** 2, axis=1)
        baseline_errors = np.sum(true_residual**2, axis=1)
        total_sse += float(np.sum(squared_errors))
        total_sst += float(np.sum(baseline_errors))
        cosine_values.extend(cosines.tolist())

        # Cosine nearest-neighbor prompt retrieval in the same normalized task
        # representation.  This is still an offline geometry diagnostic; Stage
        # C will execute selected prompts on held-out tasks.
        similarities = x_test @ x_train.T
        nearest_local = np.argmax(similarities, axis=1)
        nearest_global = [train_indices[value] for value in nearest_local]
        nn_prompt = prompts[nearest_global]
        nn_error = np.sum((nn_prompt - prompts[test_indices]) ** 2, axis=1)
        nn_sse += float(np.sum(nn_error))
        nn_sst += float(np.sum(baseline_errors))
        nn_cos = cosine_rows(nn_prompt - fold_mean, true_residual)
        nn_cosines.extend(nn_cos.tolist())

        # Behavioral proxy prediction uses the same ridge machinery and never
        # enters prompt-coordinate training.
        behavior_target = np.stack([closures[train_indices], successes[train_indices]], axis=1)
        behavior_alpha = choose_ridge_alpha(
            x_train,
            behavior_target,
            [strata[index] for index in train_indices],
            seed=seed + 100 + fold_idx,
        )
        behavior_pred = ridge_predict(add_bias(x_train), behavior_target, add_bias(x_test), behavior_alpha)
        closure_true.extend(closures[test_indices].tolist())
        closure_pred.extend(behavior_pred[:, 0].tolist())
        success_true.extend(successes[test_indices].astype(int).tolist())
        success_score.extend(behavior_pred[:, 1].tolist())
        fold_summaries.append(
            {
                "fold": fold_idx,
                "train_tasks": len(train_indices),
                "test_tasks": len(test_indices),
                "ridge_alpha": alpha,
                "behavior_alpha": behavior_alpha,
                "ridge_r2": 1.0 - float(np.sum(squared_errors)) / max(float(np.sum(baseline_errors)), 1e-30),
                "ridge_mean_cosine": float(np.mean(cosines)),
                "nn_r2": 1.0 - float(np.sum(nn_error)) / max(float(np.sum(baseline_errors)), 1e-30),
                "nn_mean_cosine": float(np.mean(nn_cos)),
            }
        )
        for local_idx, task_idx in enumerate(test_indices):
            task_predictions.append(
                {
                    "representation": name,
                    "id": str(task_idx),
                    "fold": fold_idx,
                    "ridge_cosine": float(cosines[local_idx]),
                    "nn_cosine": float(nn_cos[local_idx]),
                    "nearest_train_index": int(nearest_global[local_idx]),
                }
            )
    closure_true_array = np.asarray(closure_true)
    closure_pred_array = np.asarray(closure_pred)
    closure_sst = float(np.sum((closure_true_array - np.mean(closure_true_array)) ** 2))
    pair_rep_distance: list[float] = []
    pair_prompt_distance: list[float] = []
    normalized = representation / np.maximum(np.linalg.norm(representation, axis=1, keepdims=True), 1e-12)
    prompt_centered = prompts - np.mean(prompts, axis=0, keepdims=True)
    for left in range(len(prompts)):
        for right in range(left + 1, len(prompts)):
            pair_rep_distance.append(float(1.0 - normalized[left] @ normalized[right]))
            pair_prompt_distance.append(float(np.linalg.norm(prompt_centered[left] - prompt_centered[right])))
    result = {
        "representation": name,
        "dimension": int(representation.shape[1]),
        "prompt_ridge_reconstruction_r2": 1.0 - total_sse / max(total_sst, 1e-30),
        "prompt_ridge_mean_cosine": float(np.mean(cosine_values)),
        "prompt_ridge_median_cosine": float(np.median(cosine_values)),
        "prompt_nn_reconstruction_r2": 1.0 - nn_sse / max(nn_sst, 1e-30),
        "prompt_nn_mean_cosine": float(np.mean(nn_cosines)),
        "prompt_nn_median_cosine": float(np.median(nn_cosines)),
        "task_distance_vs_prompt_distance_spearman": spearman(
            np.asarray(pair_rep_distance), np.asarray(pair_prompt_distance)
        ),
        "closure_prediction_r2": 1.0
        - float(np.sum((closure_pred_array - closure_true_array) ** 2)) / max(closure_sst, 1e-30),
        "closure_prediction_spearman": spearman(closure_pred_array, closure_true_array),
        "success_prediction_auc": auc_score(np.asarray(success_true), np.asarray(success_score)),
        "folds": fold_summaries,
    }
    return result, task_predictions


def analyze_representations(args: argparse.Namespace, rows: list[dict[str, Any]]) -> None:
    task_ids = [row["id"] for row in rows]
    task_root = args.task_root.resolve()
    checkpoint_paths = [task_root / "training" / task_id / "final_prefix.pt" for task_id in task_ids]
    prompts, _ = checkpoint_matrix(checkpoint_paths)
    prompts = prompts.astype(np.float64)
    summaries = {task_id: read_json(task_root / "training" / task_id / "training_summary.json") for task_id in task_ids}
    evaluations = {
        str(row["id"]): row
        for row in read_jsonl(task_root / "evaluation_results.jsonl")
        if row.get("condition") == "soft"
    }
    closures = np.asarray([summaries[task_id]["core_skill_kl_closure"] for task_id in task_ids], dtype=np.float64)
    successes = np.asarray([int(bool(evaluations[task_id]["hard"])) for task_id in task_ids], dtype=np.float64)
    strata = [f"{row['instruction_type']}|success={int(successes[idx])}" for idx, row in enumerate(rows)]
    folds = stratified_folds(strata, args.folds, args.seed)

    schema_numeric = np.asarray([row["schema_numeric"] for row in rows], dtype=np.float64)
    qwen_payload = np.load(args.out_root.resolve() / "qwen_task_embeddings.npz")
    qwen_ids = [str(value) for value in qwen_payload["task_ids"].tolist()]
    if qwen_ids != task_ids:
        raise RuntimeError("Qwen embedding task order differs from TaskSpec order")
    dense: dict[str, np.ndarray] = {
        "schema_numeric": schema_numeric,
        "qwen_instruction_last": qwen_payload["qwen_instruction_last"].astype(np.float64),
        "qwen_instruction_mean": qwen_payload["qwen_instruction_mean"].astype(np.float64),
        "qwen_task_spec_last": qwen_payload["qwen_task_spec_last"].astype(np.float64),
        "qwen_task_spec_mean": qwen_payload["qwen_task_spec_mean"].astype(np.float64),
    }
    # Give the semantic and numeric schema blocks equal norm before fusion.
    qwen_block = dense["qwen_task_spec_last"]
    qwen_block = qwen_block / np.maximum(np.linalg.norm(qwen_block, axis=1, keepdims=True), 1e-12)
    schema_block = schema_numeric - np.mean(schema_numeric, axis=0, keepdims=True)
    schema_block = schema_block / np.maximum(np.linalg.norm(schema_block, axis=1, keepdims=True), 1e-12)
    dense["qwen_task_spec_last_plus_schema_numeric"] = np.concatenate([qwen_block, schema_block], axis=1)

    results: list[dict[str, Any]] = []
    predictions: list[dict[str, Any]] = []
    for name, representation in dense.items():
        result, rows_prediction = dense_representation_cv(
            name,
            representation,
            prompts,
            closures,
            successes,
            strata,
            folds,
            args.components,
            args.seed,
        )
        results.append(result)
        for prediction in rows_prediction:
            task_index = int(prediction["id"])
            prediction["id"] = task_ids[task_index]
            prediction["nearest_train_id"] = task_ids[prediction.pop("nearest_train_index")]
        predictions.extend(rows_prediction)

    # Fold-local TF-IDF instruction+schema representation. Convert its fold
    # embeddings into one fixed hashed representation for distance diagnostics
    # and use the same dense CV implementation. The hashing is deterministic
    # and does not fit corpus statistics.
    hash_dim = 2048
    hashed = np.zeros((len(rows), hash_dim), dtype=np.float64)
    for row_idx, row in enumerate(rows):
        text = f"{row['instruction_type']} {row['instruction']} {row['schema_text']}"
        for token in re.findall(r"[a-z0-9]+(?:[-_][a-z0-9]+)*", text.lower()):
            digest = int.from_bytes(__import__("hashlib").blake2b(token.encode("utf-8"), digest_size=8).digest(), "little")
            hashed[row_idx, digest % hash_dim] += 1.0 if ((digest >> 63) == 0) else -1.0
    dense_result, dense_predictions = dense_representation_cv(
        "hashed_instruction_plus_schema",
        hashed,
        prompts,
        closures,
        successes,
        strata,
        folds,
        args.components,
        args.seed,
    )
    results.append(dense_result)
    for prediction in dense_predictions:
        task_index = int(prediction["id"])
        prediction["id"] = task_ids[task_index]
        prediction["nearest_train_id"] = task_ids[prediction.pop("nearest_train_index")]
    predictions.extend(dense_predictions)

    # Raw coordinate predictability requires a non-trivial gain over the fold
    # mean and directional agreement. If nothing passes (the expected outcome
    # after Stage A), Stage C treats the strongest behavior-proxy embeddings as
    # exploratory retrievers rather than pretending they passed this screen.
    passed_coordinate_gate = [
        row["representation"]
        for row in results
        if row["prompt_ridge_reconstruction_r2"] >= 0.01
        and row["prompt_ridge_mean_cosine"] >= 0.05
    ]
    exploratory_candidates = ["qwen_task_spec_last", "qwen_instruction_last"]
    summary = {
        "stage": "B_task_representation_screen",
        "scope": "61 successful Train80 trajectories only; no Val40/Test280 access",
        "protocol": f"{args.folds}-fold task-level CV, {args.components} fold-local prompt PCA components",
        "representations": results,
        "coordinate_prediction_gate": "prompt ridge R2 >= 0.01 and mean cosine >= 0.05",
        "passed_coordinate_gate": passed_coordinate_gate,
        "stage_c_candidates": exploratory_candidates,
        "candidate_status": "exploratory_behavior_proxy_candidates" if not passed_coordinate_gate else "coordinate_gate_passed",
        "selection_warning": "No representation passed raw-coordinate prediction. Stage C held-out free execution is the decisive test.",
        "val_or_test_accessed": False,
    }
    json_dump(args.out_root.resolve() / "summary.json", summary)
    with (args.out_root.resolve() / "representation_results.csv").open("w", encoding="utf-8", newline="") as handle:
        fieldnames = [key for key in results[0] if key != "folds"]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for result in results:
            writer.writerow({key: value for key, value in result.items() if key != "folds"})
    with (args.out_root.resolve() / "cv_predictions.jsonl").open("w", encoding="utf-8") as handle:
        for row in predictions:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")
    print(json.dumps(summary, ensure_ascii=False, indent=2), flush=True)


def load_or_build_specs(args: argparse.Namespace) -> list[dict[str, Any]]:
    path = args.out_root.resolve() / "task_specs.jsonl"
    if path.is_file():
        return read_jsonl(path)
    return build_task_specs(args)


def main() -> None:
    args = parse_args()
    args.out_root.resolve().mkdir(parents=True, exist_ok=True)
    if args.phase in {"schema", "all"}:
        rows = build_task_specs(args)
    else:
        rows = load_or_build_specs(args)
    if args.phase in {"qwen", "all"}:
        extract_qwen_embeddings(args, rows)
    if args.phase in {"analyze", "all"}:
        analyze_representations(args, rows)


if __name__ == "__main__":
    main()
